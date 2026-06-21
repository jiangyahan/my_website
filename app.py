from flask import Flask, request, render_template_string, send_file, session
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import os
import re
import tempfile
import uuid

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this'

TEMP_DIR = os.path.join(tempfile.gettempdir(), 'purchase_app')
os.makedirs(TEMP_DIR, exist_ok=True)

# =================== 公用 HTML 模板 ===================
HTML_PAGE = '''
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>采购数据处理助手</title>
    <style>
        body { font-family: Arial; padding: 20px; background: #f5f7fa; }
        .container { max-width: 950px; margin: auto; }
        h1 { color: #2c3e50; }
        .function-box { background: white; padding: 25px; margin: 20px 0; border-radius: 10px; box-shadow: 0 0 8px rgba(0,0,0,0.1); }
        .function-box h2 { margin-top: 0; color: #34495e; border-bottom: 2px solid #3498db; padding-bottom: 8px; }
        label { font-weight: bold; display: block; margin: 10px 0 4px; }
        input[type="file"], input[type="number"], input[type="text"] { width: 100%; padding: 8px; margin: 5px 0 15px; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box; }
        textarea { width: 100%; padding: 8px; margin: 5px 0 15px; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box; font-family: monospace; min-height: 150px; }
        button { background: #3498db; color: white; border: none; padding: 10px 25px; border-radius: 5px; cursor: pointer; font-size: 15px; }
        button:hover { background: #2980b9; }
        .info { color: #7f8c8d; font-size: 14px; }
        .result { margin-top: 15px; padding: 15px; background: #e8f8f5; border-left: 4px solid #1abc9c; }
        .download-btn { display: inline-block; background: #2ecc71; color: white; padding: 10px 25px; text-decoration: none; border-radius: 5px; margin-top: 8px; }
        hr { margin: 30px 0; }
        .error { background: #fdedec; border-left-color: #e74c3c; }
        .example { background: #f8f9fa; padding: 10px; border-radius: 4px; font-size: 13px; border: 1px dashed #ccc; }
        .example code { white-space: pre-wrap; }
    </style>
</head>
<body>
<div class="container">
    <h1>📊 采购数据处理助手</h1>

    <!-- 功能一 -->
    <div class="function-box">
        <h2>📝 功能一：粘贴文本 → 自动解析 → 写入模板文件</h2>
        <form action="/merge" method="post" enctype="multipart/form-data">
            <label>📋 请粘贴您的采购数据</label>
            <textarea name="text_data" placeholder="例如：&#10;1、油性沥青漆 48桶&#10;到货日期：6月16送达我司" required></textarea>
            <div class="example">
                <strong>📌 支持格式：</strong><br>
                <code>1、油性沥青漆 48桶</code><br>
                <code>地板钉M8*45：482000件</code><br>
                <code>抽芯铆钉*封闭型 Ф4.8*8 支 1000 2026/6/17</code>
            </div>
            <label>📤 选择目标模板文件（B）</label>
            <input type="file" name="file_b" accept=".xlsx,.xls" required>
            <label>✏️ 请输入新工作表的名称</label>
            <input type="text" name="sheet_name" placeholder="例如：采购清单" required>
            <button type="submit">🚀 生成并下载</button>
        </form>
        {% if merge_result %}
        <div class="result {% if merge_result.error %}error{% endif %}">
            <p><strong>{{ merge_result.message }}</strong></p>
            {% if merge_result.sheet_name %}
            <p>新工作表：<strong>{{ merge_result.sheet_name }}</strong></p>
            <p>处理行数：<strong>{{ merge_result.row_count }}</strong></p>
            <p>列名：<strong>{{ merge_result.columns }}</strong></p>
            <a href="{{ merge_result.download_url }}" class="download-btn">⬇️ 下载修改后的 B 文件</a>
            {% endif %}
        </div>
        {% endif %}
    </div>

    <hr>

    <!-- 功能二 -->
    <div class="function-box">
        <h2>📦 功能二：采购清单汇总（A × 台数 → B 新增“总数”）</h2>
        <form action="/purchase" method="post" enctype="multipart/form-data">
            <label>📤 采购清单文件（A）</label>
            <input type="file" name="file_a" accept=".xlsx,.xls" required>
            <label>📤 目标模板文件（B）</label>
            <input type="file" name="file_b" accept=".xlsx,.xls" required>
            <label>🔢 输入采购台数</label>
            <input type="number" name="qty_total" step="any" value="1" required>
            <button type="submit">🚀 生成汇总</button>
        </form>
        {% if purchase_result %}
        <div class="result {% if purchase_result.error %}error{% endif %}">
            <p><strong>{{ purchase_result.message }}</strong></p>
            {% if purchase_result.sheet_name %}
            <p>新工作表：<strong>{{ purchase_result.sheet_name }}</strong></p>
            <p>处理行数：<strong>{{ purchase_result.row_count }}</strong></p>
            <a href="{{ purchase_result.download_url }}" class="download-btn">⬇️ 下载修改后的 B 文件</a>
            {% endif %}
        </div>
        {% endif %}
    </div>
</div>
</body>
</html>
'''

# =================== 免费解析器 ===================
def parse_text_to_dataframe(text):
    lines = text.strip().split('\n')
    lines = [line.strip() for line in lines if line.strip()]
    if not lines:
        return None, None, "文本为空"

    # 提取到货日期
    delivery_date = None
    date_patterns = [
        r'到货日期[：:]\s*(\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}月\d{1,2}日|\d{1,2}[-/]\d{1,2}[-/]\d{4}|\d{4}年\d{1,2}月\d{1,2}日)',
        r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
        r'(\d{1,2}月\d{1,2}日)',
    ]
    for line in lines:
        for pattern in date_patterns:
            match = re.search(pattern, line)
            if match:
                delivery_date = match.group(1)
                break
        if delivery_date:
            break

    unit_keywords = ['个', '件', '箱', '台', '套', 'kg', '克', '千克', '吨', '米', '厘米', '毫米', '升', '毫升', '包', '袋', '桶', '瓶', '盒', '卷', '片', '块', '条', '根', '支', '把', '双', '对', '副', '组', '批']
    material_rows = []

    for line in lines:
        if any(kw in line for kw in ['联系人', '电话', '订单号', '备注', '送货单', '送达', '收货', '经理', '您好']):
            continue
        if not re.search(r'\d', line):
            continue

        cleaned = re.sub(r'^\s*\d+[、.。)\]]\s*', '', line)

        name = None
        spec = None
        qty = None
        unit = None

        # 模式A：名称:数量单位
        match = re.search(r'(.+?)[：:]\s*(\d+)\s*([^\d\s]*)', cleaned)
        if match:
            name = match.group(1).strip()
            qty = float(match.group(2))
            unit_candidate = match.group(3).strip()
            if unit_candidate and any(uk in unit_candidate for uk in unit_keywords):
                unit = unit_candidate
            if '*' in name or '×' in name:
                parts = re.split(r'[*×]', name, maxsplit=1)
                if len(parts) == 2:
                    name = parts[0].strip()
                    spec = parts[1].strip()

        # 模式B：名称 数量单位
        if name is None:
            match = re.search(r'(.+?)\s+(\d+)\s*([^\d\s]*)', cleaned)
            if match:
                name = match.group(1).strip()
                qty = float(match.group(2))
                unit_candidate = match.group(3).strip()
                if unit_candidate and any(uk in unit_candidate for uk in unit_keywords):
                    unit = unit_candidate
                if '*' in name or '×' in name:
                    parts = re.split(r'[*×]', name, maxsplit=1)
                    if len(parts) == 2:
                        name = parts[0].strip()
                        spec = parts[1].strip()

        # 模式C：空格分隔多字段
        if name is None:
            parts = re.split(r'\s+', cleaned)
            for i, part in enumerate(parts):
                if re.match(r'^\d+$', part):
                    qty = float(part)
                    if i > 0 and any(uk in parts[i-1] for uk in unit_keywords):
                        unit = parts[i-1]
                    elif i+1 < len(parts) and any(uk in parts[i+1] for uk in unit_keywords):
                        unit = parts[i+1]
                    name_candidates = parts[:i]
                    if name_candidates:
                        full_name = ' '.join(name_candidates)
                        if '*' in full_name or '×' in full_name:
                            parts_spec = re.split(r'[*×]', full_name, maxsplit=1)
                            if len(parts_spec) == 2:
                                name = parts_spec[0].strip()
                                spec = parts_spec[1].strip()
                            else:
                                name = full_name
                        else:
                            name = full_name
                    else:
                        name = '材料'
                    break

        # 模式D：兜底
        if name is None:
            nums = re.findall(r'(\d+)', cleaned)
            if nums:
                qty = float(nums[0])
                for uk in unit_keywords:
                    if uk in cleaned:
                        unit = uk
                        break
                name = re.sub(r'\d+', '', cleaned)
                for uk in unit_keywords:
                    name = name.replace(uk, '')
                name = name.strip()
                if not name:
                    name = '材料'
                if '*' in name or '×' in name:
                    parts = re.split(r'[*×]', name, maxsplit=1)
                    if len(parts) == 2:
                        name = parts[0].strip()
                        spec = parts[1].strip()

        if name and qty is not None:
            if spec is None and name and ('*' in name or '×' in name):
                parts = re.split(r'[*×]', name, maxsplit=1)
                if len(parts) == 2:
                    name = parts[0].strip()
                    spec = parts[1].strip()
            if unit is None:
                for uk in unit_keywords:
                    if uk in cleaned or uk in name:
                        unit = uk
                        if uk in name:
                            name = name.replace(uk, '').strip()
                        break
                if unit is None:
                    unit = ''
            material_rows.append([name, spec or '', unit, qty, delivery_date or ''])

    if not material_rows:
        return None, None, "未能提取到任何材料数据，请检查文本格式。"

    df = pd.DataFrame(material_rows, columns=['种类', '规格', '单位', '数量', '到货日期'])
    df['数量'] = pd.to_numeric(df['数量'], errors='coerce')
    return df, df.columns.tolist(), None

# =================== 路由 ===================
@app.route('/')
def index():
    return render_template_string(HTML_PAGE)

@app.route('/merge', methods=['POST'])
def merge_files():
    text_data = request.form.get('text_data', '')
    file_b = request.files['file_b']
    sheet_name = request.form.get('sheet_name', '').strip()

    if not text_data:
        return render_template_string(HTML_PAGE, merge_result={'error': True, 'message': '❌ 错误：请粘贴数据。'})
    if not sheet_name:
        return render_template_string(HTML_PAGE, merge_result={'error': True, 'message': '❌ 错误：请填写新工作表的名称。'})
    if not file_b or file_b.filename == '':
        return render_template_string(HTML_PAGE, merge_result={'error': True, 'message': '❌ 错误：请选择目标模板文件（B）。'})

    try:
        df, columns, error = parse_text_to_dataframe(text_data)
        if error:
            return render_template_string(HTML_PAGE, merge_result={'error': True, 'message': f'❌ 解析失败：{error}'})
        if df.empty:
            return render_template_string(HTML_PAGE, merge_result={'error': True, 'message': '❌ 提取结果为空，请检查文本内容。'})

        b_bytes = file_b.read()
        b_io = BytesIO(b_bytes)
        wb = load_workbook(b_io)

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        with pd.ExcelWriter(b_io, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        temp_path = os.path.join(TEMP_DIR, f"merge_{uuid.uuid4().hex}.xlsx")
        wb.save(temp_path)

        session['merge_file_path'] = temp_path
        session['merge_filename'] = file_b.filename

        col_names = '、'.join(df.columns.tolist())
        return render_template_string(HTML_PAGE, merge_result={
            'message': f'🎉 解析成功！共识别 {len(df)} 行数据。',
            'sheet_name': sheet_name,
            'row_count': len(df),
            'columns': col_names,
            'download_url': '/download_merge'
        })
    except Exception as e:
        return render_template_string(HTML_PAGE, merge_result={'error': True, 'message': f'❌ 处理出错：{str(e)}'})

@app.route('/download_merge')
def download_merge():
    file_path = session.get('merge_file_path')
    if not file_path or not os.path.exists(file_path):
        return "文件不存在或已过期，请重新执行。"
    filename = session.get('merge_filename', '结果.xlsx')
    response = send_file(file_path, download_name=f"修改后_{filename}", as_attachment=True)
    @response.call_on_close
    def cleanup():
        try:
            os.remove(file_path)
        except:
            pass
    session.pop('merge_file_path', None)
    return response

# =================== 功能二辅助函数 ===================
def find_header_row(ws, max_rows=15):
    typical_keywords = ["图号", "名称", "规格", "材质", "厚度", "宽度", "长度", "备注"]
    for row in range(1, min(max_rows, ws.max_row) + 1):
        row_values = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str):
                row_values.append(val.strip())
        has_qty = any("数量" in v for v in row_values)
        if has_qty:
            has_other = any(any(kw in v for kw in typical_keywords) for v in row_values)
            if has_other:
                return row
    return 1

def find_quantity_column(headers):
    for idx, header in enumerate(headers):
        if header and isinstance(header, str) and "数量" in header:
            return idx
    return None

def get_number_from_cell_value(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        if val.startswith('='):
            try:
                expr = val[1:].strip()
                if re.match(r'^[\d+\-*/()\s.]+$', expr):
                    result = eval(expr)
                    return float(result)
                else:
                    nums = re.findall(r'[\d.]+', val)
                    return float(nums[0]) if nums else None
            except:
                nums = re.findall(r'[\d.]+', val)
                return float(nums[0]) if nums else None
        else:
            nums = re.findall(r'[\d.]+', val)
            return float(nums[0]) if nums else None
    return None

@app.route('/purchase', methods=['POST'])
def purchase_summary():
    file_a = request.files['file_a']
    file_b = request.files['file_b']
    qty_total = float(request.form['qty_total'])

    try:
        wb_a = load_workbook(BytesIO(file_a.read()), data_only=False)
        ws_a = wb_a.active

        header_row = find_header_row(ws_a)
        headers = []
        for col in range(1, ws_a.max_column + 1):
            val = ws_a.cell(row=header_row, column=col).value
            headers.append(val)

        qty_col_idx = find_quantity_column(headers)
        if qty_col_idx is None:
            return render_template_string(HTML_PAGE, purchase_result={'error': True, 'message': f'❌ 错误：在表头行（第{header_row}行）未找到“数量”列。'})

        data_rows = []
        for row in range(header_row + 1, ws_a.max_row + 1):
            row_data = {}
            for col in range(1, ws_a.max_column + 1):
                row_data[col] = ws_a.cell(row=row, column=col).value
            if all(v is None for v in row_data.values()):
                continue
            data_rows.append(row_data)

        totals = []
        for row_data in data_rows:
            qty_val = row_data.get(qty_col_idx + 1)
            qty_num = get_number_from_cell_value(qty_val)
            total = qty_num * qty_total if qty_num is not None else 0
            totals.append(total)

        b_bytes = file_b.read()
        b_io = BytesIO(b_bytes)
        wb_b = load_workbook(b_io)

        sheet_name = os.path.splitext(file_a.filename)[0][:31]
        if sheet_name in wb_b.sheetnames:
            del wb_b[sheet_name]

        out_ws = wb_b.create_sheet(title=sheet_name)

        new_headers = headers + ["总数"]
        for col, header in enumerate(new_headers, start=1):
            out_ws.cell(row=1, column=col, value=header)

        for r_idx, row_data in enumerate(data_rows, start=2):
            for col in range(1, ws_a.max_column + 1):
                val = row_data.get(col)
                out_ws.cell(row=r_idx, column=col, value=val)
            total_col = len(new_headers)
            out_ws.cell(row=r_idx, column=total_col, value=totals[r_idx - 2])

        temp_path = os.path.join(TEMP_DIR, f"purchase_{uuid.uuid4().hex}.xlsx")
        wb_b.save(temp_path)

        session['purchase_file_path'] = temp_path
        session['purchase_filename'] = file_b.filename

        return render_template_string(HTML_PAGE, purchase_result={
            'message': f'🎉 汇总成功！采购台数：{qty_total}',
            'sheet_name': sheet_name,
            'row_count': len(data_rows),
            'download_url': '/download_purchase'
        })
    except Exception as e:
        return render_template_string(HTML_PAGE, purchase_result={'error': True, 'message': f'❌ 处理出错：{str(e)}'})

@app.route('/download_purchase')
def download_purchase():
    file_path = session.get('purchase_file_path')
    if not file_path or not os.path.exists(file_path):
        return "文件不存在或已过期，请重新执行汇总。"
    filename = session.get('purchase_filename', '采购汇总.xlsx')
    response = send_file(file_path, download_name=f"修改后_{filename}", as_attachment=True)
    @response.call_on_close
    def cleanup():
        try:
            os.remove(file_path)
        except:
            pass
    session.pop('purchase_file_path', None)
    return response

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)